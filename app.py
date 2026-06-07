# app.py - نسخة نهائية مُحسَّنة للإنتاج (Production‑Ready)
import os, io, csv, re, hashlib
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for, send_from_directory
from database import init_db, get_db
from functools import wraps
import openpyxl
from fpdf import FPDF

app = Flask(__name__)
app.config.from_pyfile('config.py')
app.secret_key = app.config['SECRET_KEY']
init_db()

# ------------------- إعدادات الجلسة (أمان) -------------------
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',
)

# ------------------- تسجيل الدخول (مُحسَّن – لا يعتمد على X-Requested-With) -------------------
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            # إذا كان المسار يبدأ بـ /api/ فهو طلب JSON → نُرجع 401
            if request.path.startswith('/api/'):
                return jsonify({'error': 'Unauthorized', 'message': 'يرجى تسجيل الدخول'}), 401
            # وإلا نُعيد توجيه إلى صفحة الدخول
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return decorated

@app.route('/login', methods=['GET', 'POST'])
def login_page():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        pwd_hash = hashlib.sha256(password.encode()).hexdigest()
        db = get_db()
        user = db.execute("SELECT * FROM users WHERE username=? AND password_hash=?", (username, pwd_hash)).fetchone()
        db.close()
        if user:
            session['user_id'] = user['id']
            session['username'] = user['username']
            return redirect(url_for('dashboard'))
        return render_template('login.html', error='بيانات خاطئة')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login_page'))

# ------------------- PWA & Service Worker -------------------
@app.route('/manifest.json')
def manifest():
    return send_from_directory('static', 'manifest.json')

@app.route('/sw.js')
def service_worker():
    return send_from_directory('static', 'sw.js', mimetype='application/javascript')

# ------------------- أدوات مساعدة -------------------
def log_activity(action, entity_type, entity_id=None, details=''):
    db = get_db()
    db.execute("INSERT INTO activity_logs (user_id, action, entity_type, entity_id, details) VALUES (?,?,?,?,?)",
               (session.get('user_id', 1), action, entity_type, entity_id, details))
    db.commit()
    db.close()

def validate_required(data, fields):
    errs = []
    for f in fields:
        if not data.get(f) or str(data.get(f)).strip() == '':
            errs.append(f'{f} مطلوب')
    return errs

def is_number(v):
    try: float(v); return True
    except: return False

def validate_number(v, name):
    if not is_number(v): return f'{name} يجب أن يكون رقماً'
    return None

def validate_phone(p):
    if p and not re.match(r'^[\d\-\+ ]+$', p): return 'رقم هاتف غير صحيح'
    return None

def validate_percent(v):
    try:
        p = float(v)
        if 0 <= p <= 100: return None
        return 'النسبة بين 0 و 100'
    except:
        return 'النسبة غير صحيحة'

# ------------------- الصفحات -------------------
@app.route('/')
@login_required
def dashboard(): return render_template('dashboard.html')
@app.route('/customers')
@login_required
def customers(): return render_template('customers.html')
@app.route('/products')
@login_required
def products(): return render_template('products.html')
@app.route('/orders')
@login_required
def orders(): return render_template('orders.html')
@app.route('/suppliers')
@login_required
def suppliers(): return render_template('suppliers.html')
@app.route('/purchases')
@login_required
def purchases(): return render_template('purchases.html')
@app.route('/payments')
@login_required
def payments(): return render_template('payments.html')
@app.route('/commissions')
@login_required
def commissions(): return render_template('commissions.html')
@app.route('/reports')
@login_required
def reports(): return render_template('reports.html')
@app.route('/activity-logs')
@login_required
def activity_logs(): return render_template('activity_logs.html')
@app.route('/settings')
@login_required
def settings(): return render_template('settings.html')

# ==================== REST API ====================
# --- العملاء ---
@app.route('/api/customers', methods=['GET'])
@login_required
def api_customers():
    search = request.args.get('search','')
    page = int(request.args.get('page',1))
    per_page = int(request.args.get('per_page',20))
    sort_by = request.args.get('sort_by','id')
    order = request.args.get('order','asc')
    allowed = ['id','name','phone','created_at']
    if sort_by not in allowed: sort_by='id'
    order_clause = f"ORDER BY {sort_by} {order}" if order in ('asc','desc') else ''
    where = ''
    params = []
    if search:
        where = 'WHERE name LIKE ? OR phone LIKE ?'
        params = [f'%{search}%', f'%{search}%']
    db = get_db()
    total = db.execute(f"SELECT COUNT(*) FROM customers {where}", params).fetchone()[0]
    rows = db.execute(f"SELECT * FROM customers {where} {order_clause} LIMIT ? OFFSET ?",
                      params+[per_page, (page-1)*per_page]).fetchall()
    db.close()
    return jsonify({'data': [dict(r) for r in rows], 'total': total})

@app.route('/api/customers', methods=['POST'])
@login_required
def api_create_customer():
    data = request.json
    errors = validate_required(data, ['name'])
    e = validate_phone(data.get('phone',''))
    if e: errors.append(e)
    if errors: return jsonify({'errors': errors}), 400
    db = get_db()
    cur = db.execute("INSERT INTO customers (name, phone, email, address, notes) VALUES (?,?,?,?,?)",
                     (data['name'], data.get('phone',''), data.get('email',''), data.get('address',''), data.get('notes','')))
    db.commit()
    cid = cur.lastrowid
    log_activity('create','customer', cid, f"Name: {data['name']}")
    db.close()
    return jsonify({'message': 'تمت الإضافة بنجاح', 'id': cid}), 201

@app.route('/api/customers/<int:cid>', methods=['PUT'])
@login_required
def api_update_customer(cid):
    data = request.json
    errors = validate_required(data, ['name'])
    e = validate_phone(data.get('phone',''))
    if e: errors.append(e)
    if errors: return jsonify({'errors': errors}), 400
    db = get_db()
    db.execute("UPDATE customers SET name=?, phone=?, email=?, address=?, notes=? WHERE id=?",
               (data['name'], data.get('phone',''), data.get('email',''), data.get('address',''), data.get('notes',''), cid))
    db.commit()
    log_activity('update','customer', cid, f"Name: {data['name']}")
    db.close()
    return jsonify({'message': 'تم التعديل بنجاح'})

@app.route('/api/customers/<int:cid>', methods=['DELETE'])
@login_required
def api_delete_customer(cid):
    db = get_db()
    orders_count = db.execute("SELECT COUNT(*) FROM orders WHERE customer_id=?", (cid,)).fetchone()[0]
    if orders_count > 0:
        db.close()
        return jsonify({'errors': ['لا يمكن حذف العميل لوجود طلبات مسجلة عليه']}), 400
    db.execute("DELETE FROM customers WHERE id=?", (cid,))
    db.commit()
    log_activity('delete','customer', cid)
    db.close()
    return jsonify({'message': 'تم الحذف بنجاح'})

# --- المنتجات ---
@app.route('/api/products', methods=['GET'])
@login_required
def api_products():
    search = request.args.get('search','')
    page = int(request.args.get('page',1))
    per_page = int(request.args.get('per_page',20))
    sort_by = request.args.get('sort_by','id')
    order = request.args.get('order','asc')
    allowed = ['id','name','buy_price','sell_price','quantity']
    if sort_by not in allowed: sort_by='id'
    order_clause = f"ORDER BY {sort_by} {order}" if order in ('asc','desc') else ''
    where = ''
    params = []
    if search:
        where = 'WHERE name LIKE ? OR description LIKE ?'
        params = [f'%{search}%', f'%{search}%']
    db = get_db()
    total = db.execute(f"SELECT COUNT(*) FROM products {where}", params).fetchone()[0]
    rows = db.execute(f"SELECT * FROM products {where} {order_clause} LIMIT ? OFFSET ?",
                      params+[per_page, (page-1)*per_page]).fetchall()
    db.close()
    return jsonify({'data': [dict(r) for r in rows], 'total': total})

@app.route('/api/products', methods=['POST'])
@login_required
def api_create_product():
    data = request.json
    errors = validate_required(data, ['name','buy_price','sell_price'])
    e1 = validate_number(data.get('buy_price'), 'سعر الشراء')
    e2 = validate_number(data.get('sell_price'), 'سعر البيع')
    if e1: errors.append(e1)
    if e2: errors.append(e2)
    if errors: return jsonify({'errors': errors}), 400
    db = get_db()
    cur = db.execute("INSERT INTO products (name, description, buy_price, sell_price, quantity) VALUES (?,?,?,?,?)",
                     (data['name'], data.get('description',''), float(data['buy_price']), float(data['sell_price']), int(data.get('quantity',0))))
    db.commit()
    pid = cur.lastrowid
    log_activity('create','product', pid, f"Name: {data['name']}")
    db.close()
    return jsonify({'message': 'تمت الإضافة بنجاح', 'id': pid}), 201

@app.route('/api/products/<int:pid>', methods=['PUT'])
@login_required
def api_update_product(pid):
    data = request.json
    errors = validate_required(data, ['name','buy_price','sell_price'])
    e1 = validate_number(data.get('buy_price'), 'سعر الشراء')
    e2 = validate_number(data.get('sell_price'), 'سعر البيع')
    if e1: errors.append(e1)
    if e2: errors.append(e2)
    if errors: return jsonify({'errors': errors}), 400
    db = get_db()
    db.execute("UPDATE products SET name=?, description=?, buy_price=?, sell_price=?, quantity=? WHERE id=?",
               (data['name'], data.get('description',''), float(data['buy_price']), float(data['sell_price']), int(data.get('quantity',0)), pid))
    db.commit()
    log_activity('update','product', pid, f"Name: {data['name']}")
    db.close()
    return jsonify({'message': 'تم التعديل بنجاح'})

@app.route('/api/products/<int:pid>', methods=['DELETE'])
@login_required
def api_delete_product(pid):
    db = get_db()
    used_in_orders = db.execute("SELECT COUNT(*) FROM order_items WHERE product_id=?", (pid,)).fetchone()[0]
    used_in_purchases = db.execute("SELECT COUNT(*) FROM purchases WHERE product_id=?", (pid,)).fetchone()[0]
    if used_in_orders > 0 or used_in_purchases > 0:
        db.close()
        return jsonify({'errors': ['لا يمكن حذف المنتج لوجوده في سجلات طلبات أو مشتريات']}), 400
    db.execute("DELETE FROM products WHERE id=?", (pid,))
    db.commit()
    log_activity('delete','product', pid)
    db.close()
    return jsonify({'message': 'تم الحذف بنجاح'})

# --- الموردين ---
@app.route('/api/suppliers', methods=['GET'])
@login_required
def api_suppliers():
    search = request.args.get('search','')
    page = int(request.args.get('page',1))
    per_page = int(request.args.get('per_page',20))
    sort_by = request.args.get('sort_by','id')
    order = request.args.get('order','asc')
    allowed = ['id','name','phone']
    if sort_by not in allowed: sort_by='id'
    order_clause = f"ORDER BY {sort_by} {order}" if order in ('asc','desc') else ''
    where = ''
    params = []
    if search:
        where = 'WHERE name LIKE ? OR phone LIKE ?'
        params = [f'%{search}%', f'%{search}%']
    db = get_db()
    total = db.execute(f"SELECT COUNT(*) FROM suppliers {where}", params).fetchone()[0]
    rows = db.execute(f"SELECT * FROM suppliers {where} {order_clause} LIMIT ? OFFSET ?",
                      params+[per_page, (page-1)*per_page]).fetchall()
    db.close()
    return jsonify({'data': [dict(r) for r in rows], 'total': total})

@app.route('/api/suppliers', methods=['POST'])
@login_required
def api_create_supplier():
    data = request.json
    errors = validate_required(data, ['name'])
    e = validate_phone(data.get('phone',''))
    if e: errors.append(e)
    if errors: return jsonify({'errors': errors}), 400
    db = get_db()
    cur = db.execute("INSERT INTO suppliers (name, phone, email, address, notes) VALUES (?,?,?,?,?)",
                     (data['name'], data.get('phone',''), data.get('email',''), data.get('address',''), data.get('notes','')))
    db.commit()
    sid = cur.lastrowid
    log_activity('create','supplier', sid, f"Name: {data['name']}")
    db.close()
    return jsonify({'message': 'تمت الإضافة', 'id': sid}), 201

@app.route('/api/suppliers/<int:sid>', methods=['PUT'])
@login_required
def api_update_supplier(sid):
    data = request.json
    errors = validate_required(data, ['name'])
    e = validate_phone(data.get('phone',''))
    if e: errors.append(e)
    if errors: return jsonify({'errors': errors}), 400
    db = get_db()
    db.execute("UPDATE suppliers SET name=?, phone=?, email=?, address=?, notes=? WHERE id=?",
               (data['name'], data.get('phone',''), data.get('email',''), data.get('address',''), data.get('notes',''), sid))
    db.commit()
    log_activity('update','supplier', sid, f"Name: {data['name']}")
    db.close()
    return jsonify({'message': 'تم التعديل'})

@app.route('/api/suppliers/<int:sid>', methods=['DELETE'])
@login_required
def api_delete_supplier(sid):
    db = get_db()
    purchases_count = db.execute("SELECT COUNT(*) FROM purchases WHERE supplier_id=?", (sid,)).fetchone()[0]
    payments_count = db.execute("SELECT COUNT(*) FROM supplier_payments WHERE supplier_id=?", (sid,)).fetchone()[0]
    if purchases_count > 0 or payments_count > 0:
        db.close()
        return jsonify({'errors': ['لا يمكن حذف المورد لوجود مشتريات أو مدفوعات مرتبطة به']}), 400
    db.execute("DELETE FROM suppliers WHERE id=?", (sid,))
    db.commit()
    log_activity('delete','supplier', sid)
    db.close()
    return jsonify({'message': 'تم الحذف'})

# --- الطلبات (مع rollback عند الفشل) ---
@app.route('/api/orders', methods=['GET'])
@login_required
def api_orders():
    search = request.args.get('search','')
    status = request.args.get('status','')
    customer_id = request.args.get('customer_id','')
    date_from = request.args.get('date_from','')
    date_to = request.args.get('date_to','')
    page = int(request.args.get('page',1))
    per_page = int(request.args.get('per_page',20))
    sort_by = request.args.get('sort_by','id')
    order = request.args.get('order','desc')
    allowed = ['id','order_date','status','total_amount']
    if sort_by not in allowed: sort_by='id'
    order_clause = f"ORDER BY {sort_by} {order}" if order in ('asc','desc') else 'ORDER BY id DESC'
    where = []
    params = []
    if search:
        where.append("(customers.name LIKE ? OR orders.id LIKE ?)")
        params.extend([f'%{search}%', f'%{search}%'])
    if status:
        where.append("orders.status = ?")
        params.append(status)
    if customer_id:
        where.append("orders.customer_id = ?")
        params.append(int(customer_id))
    if date_from:
        where.append("orders.order_date >= ?")
        params.append(date_from)
    if date_to:
        where.append("orders.order_date <= ?")
        params.append(date_to + ' 23:59:59')
    where_clause = 'WHERE ' + ' AND '.join(where) if where else ''
    db = get_db()
    total = db.execute(f"SELECT COUNT(*) FROM orders JOIN customers ON orders.customer_id = customers.id {where_clause}", params).fetchone()[0]
    query = f"""
        SELECT orders.*, customers.name AS customer_name,
               (SELECT COALESCE(SUM(amount),0) FROM customer_payments WHERE order_id = orders.id) AS paid_amount
        FROM orders
        JOIN customers ON orders.customer_id = customers.id
        {where_clause}
        {order_clause}
        LIMIT ? OFFSET ?
    """
    rows = db.execute(query, params + [per_page, (page-1)*per_page]).fetchall()
    data = []
    for r in rows:
        d = dict(r)
        d['remaining'] = round(d['total_amount'] - d['paid_amount'], 2)
        if d['paid_amount'] >= d['total_amount']:
            d['payment_status'] = 'مدفوع'
        elif d['paid_amount'] > 0:
            d['payment_status'] = 'جزئي'
        else:
            d['payment_status'] = 'غير مدفوع'
        data.append(d)
    db.close()
    return jsonify({'data': data, 'total': total})

@app.route('/api/orders/<int:oid>', methods=['GET'])
@login_required
def api_get_order(oid):
    db = get_db()
    order = db.execute("SELECT * FROM orders WHERE id=?", (oid,)).fetchone()
    if not order:
        db.close()
        return jsonify({'error': 'غير موجود'}), 404
    items = db.execute("SELECT order_items.*, products.name AS product_name FROM order_items JOIN products ON order_items.product_id = products.id WHERE order_id=?", (oid,)).fetchall()
    db.close()
    return jsonify({'order': dict(order), 'items': [dict(i) for i in items]})

@app.route('/api/orders', methods=['POST'])
@login_required
def api_create_order():
    data = request.json
    errors = validate_required(data, ['customer_id', 'items'])
    if errors: return jsonify({'errors': errors}), 400
    if not isinstance(data['items'], list) or len(data['items'])==0:
        return jsonify({'errors': ['يجب إضافة منتج واحد على الأقل']}), 400
    db = get_db()
    cust = db.execute("SELECT id FROM customers WHERE id=?", (int(data['customer_id']),)).fetchone()
    if not cust:
        db.close()
        return jsonify({'errors': ['العميل غير موجود']}), 400
    total = 0.0
    for item in data['items']:
        product = db.execute("SELECT * FROM products WHERE id=?", (item['product_id'],)).fetchone()
        if not product:
            db.close()
            return jsonify({'errors': [f'المنتج {item["product_id"]} غير موجود']}), 400
        if product['quantity'] < item['quantity']:
            db.close()
            return jsonify({'errors': [f'المخزون غير كافٍ للمنتج {product["name"]}']}), 400
        total += item['quantity'] * item['unit_price']
    commission_percent = float(data.get('commission_percent', 0))
    commission_amount = round(total * commission_percent / 100, 2)
    cur = db.execute("INSERT INTO orders (customer_id, total_amount, commission_percent, commission_amount, notes) VALUES (?,?,?,?,?)",
                     (int(data['customer_id']), total, commission_percent, commission_amount, data.get('notes','')))
    order_id = cur.lastrowid
    for item in data['items']:
        db.execute("INSERT INTO order_items (order_id, product_id, quantity, unit_price) VALUES (?,?,?,?)",
                   (order_id, item['product_id'], item['quantity'], item['unit_price']))
        db.execute("UPDATE products SET quantity = quantity - ? WHERE id=?", (item['quantity'], item['product_id']))
    db.commit()
    log_activity('create','order', order_id, f"Customer ID: {data['customer_id']}, Total: {total}")
    db.close()
    return jsonify({'message': 'تم إنشاء الطلب بنجاح', 'id': order_id}), 201

@app.route('/api/orders/<int:oid>', methods=['PUT'])
@login_required
def api_update_order(oid):
    data = request.json
    errors = validate_required(data, ['customer_id', 'items'])
    if errors: return jsonify({'errors': errors}), 400
    if not isinstance(data['items'], list) or len(data['items'])==0:
        return jsonify({'errors': ['يجب إضافة منتج واحد على الأقل']}), 400
    db = get_db()
    old_order = db.execute("SELECT * FROM orders WHERE id=?", (oid,)).fetchone()
    if not old_order:
        db.close()
        return jsonify({'errors': ['الطلب غير موجود']}), 404
    try:
        # إعادة المخزون القديم
        old_items = db.execute("SELECT * FROM order_items WHERE order_id=?", (oid,)).fetchall()
        for oi in old_items:
            db.execute("UPDATE products SET quantity = quantity + ? WHERE id=?", (oi['quantity'], oi['product_id']))
        db.execute("DELETE FROM order_items WHERE order_id=?", (oid,))
        total = 0.0
        for item in data['items']:
            product = db.execute("SELECT * FROM products WHERE id=?", (item['product_id'],)).fetchone()
            if not product:
                raise Exception(f"المنتج {item['product_id']} غير موجود")
            if product['quantity'] < item['quantity']:
                raise Exception(f"المخزون غير كافٍ للمنتج {product['name']}")
            total += item['quantity'] * item['unit_price']
            db.execute("INSERT INTO order_items (order_id, product_id, quantity, unit_price) VALUES (?,?,?,?)",
                       (oid, item['product_id'], item['quantity'], item['unit_price']))
            db.execute("UPDATE products SET quantity = quantity - ? WHERE id=?", (item['quantity'], item['product_id']))
        commission_percent = float(data.get('commission_percent', old_order['commission_percent']))
        commission_amount = round(total * commission_percent / 100, 2)
        db.execute("UPDATE orders SET customer_id=?, total_amount=?, commission_percent=?, commission_amount=?, notes=? WHERE id=?",
                   (int(data['customer_id']), total, commission_percent, commission_amount, data.get('notes',''), oid))
        db.commit()
        log_activity('update','order', oid, f"New total: {total}")
    except Exception as e:
        db.rollback()
        db.close()
        return jsonify({'errors': [str(e)]}), 400
    db.close()
    return jsonify({'message': 'تم تعديل الطلب بنجاح'})

@app.route('/api/orders/<int:oid>/status', methods=['PUT'])
@login_required
def api_update_order_status(oid):
    data = request.json
    new_status = data.get('status')
    if new_status not in ['pending','confirmed','shipping','delivered','cancelled']:
        return jsonify({'errors': ['حالة غير صالحة']}), 400
    db = get_db()
    order = db.execute("SELECT * FROM orders WHERE id=?", (oid,)).fetchone()
    if not order:
        db.close()
        return jsonify({'errors': ['الطلب غير موجود']}), 404
    old_status = order['status']
    if new_status == old_status:
        db.close()
        return jsonify({'message': 'لم يحدث تغيير'})
    if new_status == 'cancelled' and old_status != 'cancelled':
        items = db.execute("SELECT * FROM order_items WHERE order_id=?", (oid,)).fetchall()
        for item in items:
            db.execute("UPDATE products SET quantity = quantity + ? WHERE id=?", (item['quantity'], item['product_id']))
    elif old_status == 'cancelled' and new_status != 'cancelled':
        items = db.execute("SELECT * FROM order_items WHERE order_id=?", (oid,)).fetchall()
        for item in items:
            prod = db.execute("SELECT quantity FROM products WHERE id=?", (item['product_id'],)).fetchone()
            if prod['quantity'] < item['quantity']:
                db.close()
                return jsonify({'errors': [f'لا يمكن إعادة تنشيط الطلب، المخزون غير كافٍ للمنتج {item["product_id"]}']}), 400
            db.execute("UPDATE products SET quantity = quantity - ? WHERE id=?", (item['quantity'], item['product_id']))
    db.execute("UPDATE orders SET status=? WHERE id=?", (new_status, oid))
    db.commit()
    log_activity('update','order', oid, f"Status changed from {old_status} to {new_status}")
    db.close()
    return jsonify({'message': 'تم تغيير الحالة'})

@app.route('/api/orders/<int:oid>', methods=['DELETE'])
@login_required
def api_delete_order(oid):
    db = get_db()
    order = db.execute("SELECT status FROM orders WHERE id=?", (oid,)).fetchone()
    if order and order['status'] != 'cancelled':
        items = db.execute("SELECT * FROM order_items WHERE order_id=?", (oid,)).fetchall()
        for item in items:
            db.execute("UPDATE products SET quantity = quantity + ? WHERE id=?", (item['quantity'], item['product_id']))
    db.execute("DELETE FROM orders WHERE id=?", (oid,))
    db.commit()
    log_activity('delete','order', oid)
    db.close()
    return jsonify({'message': 'تم حذف الطلب'})

# --- مدفوعات العملاء ---
@app.route('/api/customer-payments', methods=['GET'])
@login_required
def api_customer_payments():
    order_id = request.args.get('order_id')
    db = get_db()
    if order_id:
        rows = db.execute("SELECT * FROM customer_payments WHERE order_id=? ORDER BY payment_date DESC", (order_id,)).fetchall()
    else:
        rows = db.execute("SELECT * FROM customer_payments ORDER BY payment_date DESC LIMIT 100").fetchall()
    db.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/customer-payments', methods=['POST'])
@login_required
def api_create_customer_payment():
    data = request.json
    errors = validate_required(data, ['order_id','amount'])
    e = validate_number(data.get('amount'), 'المبلغ')
    if e: errors.append(e)
    if errors: return jsonify({'errors': errors}), 400
    db = get_db()
    order = db.execute("SELECT * FROM orders WHERE id=?", (data['order_id'],)).fetchone()
    if not order:
        db.close()
        return jsonify({'errors': ['الطلب غير موجود']}), 404
    paid = db.execute("SELECT COALESCE(SUM(amount),0) FROM customer_payments WHERE order_id=?", (data['order_id'],)).fetchone()[0]
    remaining = order['total_amount'] - paid
    if float(data['amount']) > remaining:
        db.close()
        return jsonify({'errors': ['المبلغ يتجاوز المتبقي']}), 400
    db.execute("INSERT INTO customer_payments (order_id, amount, payment_date) VALUES (?,?,?)",
               (data['order_id'], float(data['amount']), data.get('payment_date', datetime.now().strftime('%Y-%m-%d'))))
    db.commit()
    log_activity('create','customer_payment', None, f"Order #{data['order_id']} paid {data['amount']}")
    db.close()
    return jsonify({'message': 'تم تسجيل الدفعة'})

@app.route('/api/customer-payments/<int:pid>', methods=['DELETE'])
@login_required
def api_delete_customer_payment(pid):
    db = get_db()
    db.execute("DELETE FROM customer_payments WHERE id=?", (pid,))
    db.commit()
    log_activity('delete','customer_payment', pid)
    db.close()
    return jsonify({'message': 'تم حذف الدفعة'})

# --- مدفوعات الموردين ---
@app.route('/api/supplier-payments', methods=['GET'])
@login_required
def api_supplier_payments():
    supplier_id = request.args.get('supplier_id')
    db = get_db()
    if supplier_id:
        rows = db.execute("SELECT * FROM supplier_payments WHERE supplier_id=? ORDER BY payment_date DESC", (supplier_id,)).fetchall()
    else:
        rows = db.execute("SELECT * FROM supplier_payments ORDER BY payment_date DESC LIMIT 100").fetchall()
    db.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/supplier-payments', methods=['POST'])
@login_required
def api_create_supplier_payment():
    data = request.json
    errors = validate_required(data, ['supplier_id','amount'])
    e = validate_number(data.get('amount'), 'المبلغ')
    if e: errors.append(e)
    if errors: return jsonify({'errors': errors}), 400
    db = get_db()
    db.execute("INSERT INTO supplier_payments (supplier_id, purchase_id, amount, payment_date) VALUES (?,?,?,?)",
               (data['supplier_id'], data.get('purchase_id'), float(data['amount']), data.get('payment_date', datetime.now().strftime('%Y-%m-%d'))))
    db.commit()
    log_activity('create','supplier_payment', None, f"Supplier #{data['supplier_id']} paid {data['amount']}")
    db.close()
    return jsonify({'message': 'تم تسجيل الدفعة'})

# --- المشتريات ---
@app.route('/api/purchases', methods=['GET'])
@login_required
def api_purchases():
    search = request.args.get('search','')
    page = int(request.args.get('page',1))
    per_page = int(request.args.get('per_page',20))
    sort_by = request.args.get('sort_by','id')
    order = request.args.get('order','desc')
    allowed = ['id','purchase_date','quantity']
    if sort_by not in allowed: sort_by='id'
    order_clause = f"ORDER BY {sort_by} {order}" if order in ('asc','desc') else ''
    where = []
    params = []
    if search:
        where.append("(suppliers.name LIKE ? OR products.name LIKE ?)")
        params.extend([f'%{search}%', f'%{search}%'])
    where_clause = 'WHERE ' + ' AND '.join(where) if where else ''
    db = get_db()
    total = db.execute(f"SELECT COUNT(*) FROM purchases JOIN suppliers ON purchases.supplier_id=suppliers.id JOIN products ON purchases.product_id=products.id {where_clause}", params).fetchone()[0]
    query = f"""
        SELECT purchases.*, suppliers.name AS supplier_name, products.name AS product_name
        FROM purchases
        JOIN suppliers ON purchases.supplier_id=suppliers.id
        JOIN products ON purchases.product_id=products.id
        {where_clause}
        {order_clause}
        LIMIT ? OFFSET ?
    """
    rows = db.execute(query, params + [per_page, (page-1)*per_page]).fetchall()
    db.close()
    return jsonify({'data': [dict(r) for r in rows], 'total': total})

@app.route('/api/purchases', methods=['POST'])
@login_required
def api_create_purchase():
    data = request.json
    errors = validate_required(data, ['supplier_id','product_id','quantity','buy_price'])
    e = validate_number(data.get('quantity'), 'الكمية')
    if e: errors.append(e)
    e2 = validate_number(data.get('buy_price'), 'سعر الشراء')
    if e2: errors.append(e2)
    if errors: return jsonify({'errors': errors}), 400
    db = get_db()
    db.execute("INSERT INTO purchases (supplier_id, product_id, quantity, buy_price) VALUES (?,?,?,?)",
               (int(data['supplier_id']), int(data['product_id']), int(data['quantity']), float(data['buy_price'])))
    db.execute("UPDATE products SET quantity = quantity + ? WHERE id=?", (int(data['quantity']), int(data['product_id'])))
    db.commit()
    log_activity('create','purchase', None, f"Product {data['product_id']} qty {data['quantity']}")
    db.close()
    return jsonify({'message': 'تمت إضافة المشتريات'})

@app.route('/api/purchases/<int:pid>', methods=['DELETE'])
@login_required
def api_delete_purchase(pid):
    db = get_db()
    purchase = db.execute("SELECT * FROM purchases WHERE id=?", (pid,)).fetchone()
    if not purchase:
        db.close()
        return jsonify({'errors': ['المشتريات غير موجودة']}), 404
    payments_linked = db.execute("SELECT COUNT(*) FROM supplier_payments WHERE purchase_id=?", (pid,)).fetchone()[0]
    if payments_linked > 0:
        db.close()
        return jsonify({'errors': ['لا يمكن حذف المشتريات لوجود مدفوعات مرتبطة بها']}), 400
    cur_qty = db.execute("SELECT quantity FROM products WHERE id=?", (purchase['product_id'],)).fetchone()[0]
    if cur_qty < purchase['quantity']:
        db.close()
        return jsonify({'errors': ['لا يمكن حذف المشتريات، المخزون الحالي أقل من الكمية المشتراة']}), 400
    db.execute("UPDATE products SET quantity = quantity - ? WHERE id=?", (purchase['quantity'], purchase['product_id']))
    db.execute("DELETE FROM purchases WHERE id=?", (pid,))
    db.commit()
    log_activity('delete','purchase', pid)
    db.close()
    return jsonify({'message': 'تم الحذف'})

# --- العمولات (عرض) ---
@app.route('/api/commissions', methods=['GET'])
@login_required
def api_commissions():
    search = request.args.get('search','')
    page = int(request.args.get('page',1))
    per_page = int(request.args.get('per_page',20))
    where = ''
    params = []
    if search:
        where = 'WHERE orders.id LIKE ? OR customers.name LIKE ?'
        params = [f'%{search}%', f'%{search}%']
    db = get_db()
    total = db.execute(f"SELECT COUNT(*) FROM orders JOIN customers ON orders.customer_id=customers.id {where}", params).fetchone()[0]
    rows = db.execute(f"""
        SELECT orders.id, customers.name AS customer_name, orders.total_amount,
               orders.commission_percent, orders.commission_amount, orders.status
        FROM orders
        JOIN customers ON orders.customer_id=customers.id
        {where}
        ORDER BY orders.id DESC
        LIMIT ? OFFSET ?
    """, params + [per_page, (page-1)*per_page]).fetchall()
    db.close()
    return jsonify({'data': [dict(r) for r in rows], 'total': total})

# --- لوحة التحكم (إحصائيات المبيعات والعمولات للمكتملة فقط) ---
@app.route('/api/dashboard/stats')
@login_required
def api_dashboard_stats():
    db = get_db()
    customers_count = db.execute("SELECT COUNT(*) FROM customers").fetchone()[0]
    products_count = db.execute("SELECT COUNT(*) FROM products").fetchone()[0]
    orders_count = db.execute("SELECT COUNT(*) FROM orders").fetchone()[0]
    pending_orders = db.execute("SELECT COUNT(*) FROM orders WHERE status IN ('pending','confirmed','shipping')").fetchone()[0]

    total_sales = db.execute("""
        SELECT COALESCE(SUM(o.total_amount), 0)
        FROM orders o
        WHERE o.status = 'delivered'
          AND (SELECT COALESCE(SUM(cp.amount),0) FROM customer_payments cp WHERE cp.order_id = o.id) >= o.total_amount
    """).fetchone()[0]

    total_commissions = db.execute("""
        SELECT COALESCE(SUM(o.commission_amount), 0)
        FROM orders o
        WHERE o.status = 'delivered'
          AND (SELECT COALESCE(SUM(cp.amount),0) FROM customer_payments cp WHERE cp.order_id = o.id) >= o.total_amount
    """).fetchone()[0]

    low_stock = db.execute("SELECT COUNT(*) FROM products WHERE quantity <= low_stock_threshold").fetchone()[0]
    total_customer_payments = db.execute("SELECT COALESCE(SUM(amount),0) FROM customer_payments").fetchone()[0]
    total_supplier_payments = db.execute("SELECT COALESCE(SUM(amount),0) FROM supplier_payments").fetchone()[0]
    recent_orders = db.execute("""
        SELECT orders.id, customers.name, orders.total_amount, orders.status
        FROM orders JOIN customers ON orders.customer_id=customers.id
        ORDER BY orders.order_date DESC LIMIT 5
    """).fetchall()
    db.close()
    return jsonify({
        'customers': customers_count,
        'products': products_count,
        'orders': orders_count,
        'pending_orders': pending_orders,
        'total_sales': round(total_sales,2),
        'total_commissions': round(total_commissions,2),
        'low_stock': low_stock,
        'total_customer_payments': round(total_customer_payments,2),
        'total_supplier_payments': round(total_supplier_payments,2),
        'recent_orders': [dict(r) for r in recent_orders]
    })

# --- التقارير (المبيعات تشمل المكتملة فقط) ---
@app.route('/api/reports/sales')
@login_required
def api_report_sales():
    db = get_db()
    rows = db.execute("""
        SELECT
            strftime('%Y-%m', o.order_date) as month,
            COUNT(*) as count,
            SUM(CASE
                WHEN o.status = 'delivered'
                     AND (SELECT COALESCE(SUM(cp.amount),0) FROM customer_payments cp WHERE cp.order_id = o.id) >= o.total_amount
                THEN o.total_amount
                ELSE 0
            END) as total
        FROM orders o
        WHERE o.status != 'cancelled'
        GROUP BY month
        ORDER BY month DESC
    """).fetchall()
    db.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/reports/low-stock')
@login_required
def api_report_low_stock():
    db = get_db()
    rows = db.execute("SELECT * FROM products WHERE quantity <= low_stock_threshold").fetchall()
    db.close()
    return jsonify([dict(r) for r in rows])

# --- التصدير (PDF بشكل آمن) ---
@app.route('/api/export/<entity>')
@login_required
def api_export(entity):
    fmt = request.args.get('format','csv')
    if entity == 'customers':
        db = get_db()
        rows = db.execute("SELECT id, name, phone, email, address FROM customers").fetchall()
        db.close()
        data = [dict(r) for r in rows]
        columns = ['id','name','phone','email','address']
    elif entity == 'orders':
        db = get_db()
        rows = db.execute("SELECT orders.id, customers.name, orders.total_amount, orders.status, orders.commission_amount FROM orders JOIN customers ON orders.customer_id=customers.id").fetchall()
        db.close()
        data = [dict(r) for r in rows]
        columns = ['id','customer','total','status','commission']
    else:
        return 'غير مدعوم', 400
    if fmt == 'csv':
        si = io.StringIO()
        cw = csv.writer(si)
        cw.writerow(columns)
        for row in data:
            cw.writerow([row.get(col,'') for col in columns])
        output = io.BytesIO()
        output.write(si.getvalue().encode('utf-8-sig'))
        output.seek(0)
        return send_file(output, mimetype='text/csv', as_attachment=True, download_name=f'{entity}.csv')
    elif fmt == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(columns)
        for row in data:
            ws.append([row.get(col,'') for col in columns])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f'{entity}.xlsx')
    elif fmt == 'pdf':
        pdf = FPDF()
        pdf.add_page()
        pdf.add_font('Arial', '', 'arial.ttf', uni=True)
        pdf.set_font('Arial', '', 12)
        for col in columns:
            pdf.cell(40, 10, col, 1)
        pdf.ln()
        for row in data:
            for col in columns:
                pdf.cell(40, 10, str(row.get(col,'')), 1)
            pdf.ln()
        # fpdf2 2.x: output() returns bytes
        pdf_bytes = pdf.output()
        output = io.BytesIO(pdf_bytes)
        output.seek(0)
        return send_file(output, mimetype='application/pdf', as_attachment=True, download_name=f'{entity}.pdf')
    return 'صيغة غير مدعومة', 400

# --- سجل النشاط ---
@app.route('/api/activity-logs')
@login_required
def api_activity_logs():
    page = int(request.args.get('page',1))
    per_page = int(request.args.get('per_page',50))
    db = get_db()
    total = db.execute("SELECT COUNT(*) FROM activity_logs").fetchone()[0]
    rows = db.execute("SELECT * FROM activity_logs ORDER BY created_at DESC LIMIT ? OFFSET ?", (per_page, (page-1)*per_page)).fetchall()
    db.close()
    return jsonify({'data': [dict(r) for r in rows], 'total': total})

# --- الإعدادات ---
@app.route('/api/change-password', methods=['POST'])
@login_required
def api_change_password():
    data = request.json
    errors = validate_required(data, ['current_password','new_password'])
    if errors: return jsonify({'errors': errors}), 400
    pwd_hash_old = hashlib.sha256(data['current_password'].encode()).hexdigest()
    db = get_db()
    user = db.execute("SELECT * FROM users WHERE id=?", (session['user_id'],)).fetchone()
    if user['password_hash'] != pwd_hash_old:
        db.close()
        return jsonify({'errors': ['كلمة المرور الحالية غير صحيحة']}), 400
    new_hash = hashlib.sha256(data['new_password'].encode()).hexdigest()
    db.execute("UPDATE users SET password_hash=? WHERE id=?", (new_hash, session['user_id']))
    db.commit()
    db.close()
    return jsonify({'message': 'تم تغيير كلمة المرور'})

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=False)
